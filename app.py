from dotenv import load_dotenv
load_dotenv()

import os
import jwt
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file
from google.oauth2 import id_token
from google.auth.transport import requests as google_requests
from config import (ESTADOS, MEDIOS_PAGO,
                    PRECIO_LOCRO_UNITARIO, PRECIO_PASTELITO_DOCENA,
                    PRECIO_PASTELITO_MEDIA_DOCENA, PRECIO_PASTELITO_UNIDAD,
                    CODIGO_ADMIN)
import models
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from flask_cors import CORS
import traceback
from werkzeug.exceptions import HTTPException
from db import init_db_app
from flask_socketio import SocketIO

app = Flask(__name__)
CORS(app) # Habilita CORS para todas las rutas y orígenes

init_db_app(app)
socketio = SocketIO(app, cors_allowed_origins="*")

@app.errorhandler(Exception)
def handle_exception(e):
    if isinstance(e, HTTPException):
        return jsonify(error=e.description), e.code
    print(f"Error inesperado: {str(e)}")
    traceback.print_exc()
    return jsonify(error="Error interno del servidor", details=str(e)), 500


SECRET_KEY = os.environ.get('JWT_SECRET_KEY')
if not SECRET_KEY:
    raise ValueError("Falta la variable de entorno JWT_SECRET_KEY. Es requerida para la seguridad del sistema.")

def generar_token(user_id):
    payload = {
        'exp': datetime.utcnow() + timedelta(days=30),
        'iat': datetime.utcnow(),
        'sub': str(user_id)
    }
    return jwt.encode(payload, SECRET_KEY, algorithm='HS256')

def obtener_usuario_desde_token(token):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=['HS256'])
        return int(payload['sub'])
    except jwt.ExpiredSignatureError:
        return 'token_expirado'
    except jwt.InvalidTokenError:
        return 'token_invalido'

def obtener_usuario_id_opcional():
    token = None
    if 'Authorization' in request.headers:
        auth_header = request.headers['Authorization']
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    if not token:
        return None
    user_id = obtener_usuario_desde_token(token)
    if user_id in ('token_expirado', 'token_invalido'):
        return None
    return user_id

def obtener_usuario_actual():
    user_id = obtener_usuario_id_opcional()
    if not user_id:
        return None
    from db import Usuario
    return Usuario.query.get(user_id)

# ── Inicializar base de datos al arrancar ────────────────────────────────────
with app.app_context():
    models.init_db()


# ── Helpers de validación ────────────────────────────────────────────────────
def validar_pedido(data):
    """Valida los campos obligatorios de un pedido. Retorna lista de errores."""
    errores = []
    if not data.get('nombre_cliente', '').strip():
        errores.append('El nombre del cliente es obligatorio.')
    if not data.get('telefono', '').strip():
        errores.append('El teléfono es obligatorio.')
    if not data.get('direccion', '').strip():
        errores.append('La dirección es obligatoria.')
    if data.get('medio_pago') not in MEDIOS_PAGO:
        errores.append(f"Medio de pago inválido. Debe ser: {', '.join(MEDIOS_PAGO)}.")
    if data.get('tipo_entrega') not in ('envio', 'retiro'):
        errores.append("Tipo de entrega inválido.")
    
    # Validación de formato de email
    email = data.get('email', '').strip()
    if email:
        import re
        if not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
            errores.append('El formato del email es inválido.')

    # Validación de formato de teléfono
    telefono = data.get('telefono', '').strip()
    if telefono:
        import re
        if not re.match(r"^[+0-9\s()-]+$", telefono):
            errores.append('El teléfono contiene caracteres no válidos.')

    # Mitigación XSS: Bloquear caracteres < y > en campos de texto
    for field in ['nombre_cliente', 'direccion', 'notas', 'horario_entrega']:
        val = data.get(field, '')
        if val and ('<' in val or '>' in val):
            errores.append(f'El campo {field.replace("_", " ")} no puede contener los caracteres "<" o ">".')

    try:
        items = data.get('items', [])
        if not items:
            errores.append('El pedido debe tener al menos un producto.')
        else:
            for it in items:
                cant = int(it.get('cantidad', 0))
                if cant < 0:
                    errores.append('Las cantidades no pueden ser negativas.')
                    break
    except (ValueError, TypeError):
        errores.append('Las cantidades deben ser números enteros.')
    return errores


# ── Rutas de páginas ─────────────────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard'))


@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/productos')
def gestion_productos():
    return render_template('productos.html')

@app.route('/nuevo-pedido')
def nuevo_pedido():
    return render_template('nuevo_pedido.html')


@app.route('/pedidos/<int:pedido_id>/editar', methods=['GET'])
def editar_pedido_form(pedido_id):
    pedido = models.get_pedido_by_id(pedido_id)
    if not pedido:
        return render_template('dashboard.html'), 404
    return render_template('editar_pedido.html', pedido=pedido, estados=ESTADOS)


@app.route('/api/pedidos/<int:pedido_id>/editar', methods=['POST'])
def editar_pedido_submit(pedido_id):
    pedido = models.get_pedido_by_id(pedido_id)
    if not pedido:
        return jsonify({'error': 'Pedido no encontrado.'}), 404

    data = request.get_json() or request.form.to_dict()
    errores = validar_pedido(data)
    if data.get('estado') not in ESTADOS:
        errores.append(f"Estado inválido.")
    if errores:
        return jsonify({'errores': errores}), 400

    ok, monto_total = models.update_pedido(pedido_id, data)
    if not ok:
        return jsonify({'error': 'No se pudo actualizar el pedido.'}), 500
    return jsonify({'ok': True, 'monto_total': monto_total})


# ── API de productos ─────────────────────────────────────────────────────────

@app.route('/api/productos', methods=['GET'])
def listar_productos():
    solo_activos = request.args.get('activos') == 'true'
    productos = models.get_productos(solo_activos=solo_activos)
    return jsonify(productos)

@app.route('/api/productos', methods=['POST'])
def crear_producto():
    data = request.get_json()
    if not data or not data.get('nombre') or not data.get('precio'):
        return jsonify({'error': 'Faltan datos obligatorios (nombre, precio).'}), 400
    try:
        p_id = models.create_producto(data)
        return jsonify({'ok': True, 'id': p_id}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/productos/<int:prod_id>', methods=['PUT'])
def modificar_producto(prod_id):
    data = request.get_json()
    try:
        ok = models.update_producto(prod_id, data)
        if not ok:
            return jsonify({'error': 'Producto no encontrado.'}), 404
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/productos/<int:prod_id>', methods=['DELETE'])
def borrar_producto(prod_id):
    try:
        models.delete_producto(prod_id)
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'error': 'No se puede eliminar un producto que ya tiene pedidos asociados.'}), 409


# ── API de pedidos ───────────────────────────────────────────────────────────

@app.route('/api/pedidos', methods=['POST'])
def crear_pedido():
    data = request.get_json() or request.form.to_dict()
    from schemas import pedido_schema
    from marshmallow import ValidationError
    try:
        valid_data = pedido_schema.load(data)
    except ValidationError as err:
        # Formatear errores de marshmallow en una lista de strings
        errores = [f"{campo}: {msgs[0]}" for campo, msgs in err.messages.items()]
        return jsonify({'errores': errores}), 400

    user_id = obtener_usuario_id_opcional()
    if user_id:
        valid_data['usuario_id'] = user_id

    pedido_id, monto_total = models.create_pedido(valid_data)
    socketio.emit('pedidos_actualizados', {'mensaje': 'Nuevo pedido creado'})
    return jsonify({'ok': True, 'id': pedido_id, 'monto_total': monto_total}), 201


@app.route('/api/pedidos', methods=['GET'])
def listar_pedidos():
    usuario = obtener_usuario_actual()
    usuario_id_filtro = None
    if usuario and usuario.rol != 'admin':
        usuario_id_filtro = usuario.id

    estado     = request.args.get('estado') or None
    medio_pago = request.args.get('medio_pago') or None
    fecha      = request.args.get('fecha') or None
    busqueda   = request.args.get('q') or None
    tipo_entrega = request.args.get('tipo_entrega') or None
    page = request.args.get('page', type=int)
    limit = request.args.get('limit', default=30, type=int)

    pedidos    = models.get_all_pedidos(
        estado=estado, 
        medio_pago=medio_pago, 
        fecha=fecha, 
        busqueda=busqueda, 
        tipo_entrega=tipo_entrega, 
        usuario_id_filtro=usuario_id_filtro,
        page=page,
        limit=limit
    )
    return jsonify(pedidos)


@app.route('/api/pedidos/fechas', methods=['GET'])
def listar_fechas_pedidos():
    fechas = models.get_fechas_pedidos()
    return jsonify(fechas)


@app.route('/api/pedidos/mis-pedidos', methods=['GET'])
def listar_mis_pedidos():
    user_id = obtener_usuario_id_opcional()
    if not user_id:
        return jsonify({'error': 'No autorizado.'}), 401
    pedidos = models.get_pedidos_by_usuario(user_id)
    return jsonify(pedidos)


@app.route('/api/pedidos/<int:pedido_id>', methods=['GET'])
def detalle_pedido(pedido_id):
    pedido = models.get_pedido_by_id(pedido_id)
    if not pedido:
        return jsonify({'error': 'Pedido no encontrado.'}), 404
    return jsonify(pedido)


@app.route('/api/pedidos/<int:pedido_id>/pagado', methods=['PUT'])
def actualizar_pagado(pedido_id):
    data = request.get_json()
    if 'pagado' not in data:
        return jsonify({'error': 'Falta el campo pagado.'}), 400
    if models.update_pagado(pedido_id, data['pagado']):
        socketio.emit('pedidos_actualizados', {'mensaje': 'Pago actualizado'})
        return jsonify({'ok': True})
    return jsonify({'ok': True, 'pagado': pedido.pagado})

@app.route('/api/pedidos/<int:pedido_id>/direccion', methods=['PUT'])
def actualizar_direccion_pedido(pedido_id):
    pedido = models.get_pedido_by_id(pedido_id)
    if not pedido:
        return jsonify({'error': 'Pedido no encontrado'}), 404

    # Only allow updating direction if pending or confirmed
    if pedido['estado'] not in ['Pendiente', 'Confirmado']:
        return jsonify({'error': 'No se puede cambiar la dirección en este estado'}), 400

    data = request.get_json() or {}
    new_direccion = data.get('direccion', '').strip()
    
    if not new_direccion:
        return jsonify({'error': 'La dirección no puede estar vacía'}), 400
        
    if '<' in new_direccion or '>' in new_direccion:
        return jsonify({'error': 'La dirección contiene caracteres inválidos'}), 400

    # We need to update the model. I will add a helper in models.py or just use SQLAlchemy here.
    # It's better to do it cleanly.
    p_obj = models.Pedido.query.get(pedido_id)
    if p_obj:
        import bleach
        clean_dir = bleach.clean(new_direccion)
        if clean_dir != p_obj.direccion:
            p_obj.direccion = clean_dir
            p_obj.direccion_editada = True
            models.db.session.commit()
            socketio.emit('pedidos_actualizados', {'mensaje': 'Dirección actualizada'})

    return jsonify({'ok': True, 'direccion': new_direccion, 'direccion_editada': True})


@app.route('/api/pedidos/<int:pedido_id>/repartidor', methods=['PUT'])
def actualizar_repartidor(pedido_id):
    data = request.get_json()
    if 'repartidor' not in data:
        return jsonify({'error': 'Falta el campo repartidor.'}), 400
    if models.update_repartidor(pedido_id, data['repartidor']):
        socketio.emit('pedidos_actualizados', {'mensaje': 'Repartidor actualizado'})
        return jsonify({'ok': True})
    return jsonify({'error': 'Pedido no encontrado.'}), 404


@app.route('/api/pedidos/<int:pedido_id>/estado', methods=['PUT'])
def cambiar_estado(pedido_id):
    data = request.get_json()
    from schemas import cambio_estado_schema
    from marshmallow import ValidationError
    try:
        valid_data = cambio_estado_schema.load(data or {})
    except ValidationError as err:
        return jsonify({'error': 'Estado inválido.'}), 400
        
    try:
        ok = models.update_estado(pedido_id, valid_data['estado'])
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    if not ok:
        return jsonify({'error': 'Pedido no encontrado.'}), 404
    socketio.emit('pedidos_actualizados', {'mensaje': 'Estado de pedido actualizado'})
    return jsonify({'ok': True, 'estado': valid_data['estado']})


@app.route('/api/pedidos/<int:pedido_id>', methods=['DELETE'])
def eliminar_pedido(pedido_id):
    ok, motivo = models.delete_pedido(pedido_id)
    if not ok:
        if motivo == 'not_found':
            return jsonify({'error': 'Pedido no encontrado.'}), 404
        if motivo == 'not_pending':
            return jsonify({'error': 'Solo se pueden eliminar pedidos en estado Pendiente.'}), 409
    return jsonify({'ok': True})


# ── Nuevas Vistas (Agenda y Reparto) ─────────────────────────────────────────

@app.route('/agenda')
def agenda():
    return render_template('agenda.html')


@app.route('/reparto')
def reparto():
    return render_template('reparto.html')


# ── Nuevas API de Contactos y Envío ──────────────────────────────────────────

@app.route('/api/contactos', methods=['GET'])
def listar_contactos():
    fecha = request.args.get('fecha') or None
    contactos = models.get_contactos(fecha=fecha)
    return jsonify(contactos)


@app.route('/api/contactos/historial', methods=['GET'])
def historial_contacto():
    nombre = request.args.get('nombre')
    telefono = request.args.get('telefono')
    if not nombre or not telefono:
        return jsonify({'error': 'Nombre y teléfono son obligatorios.'}), 400
    historial = models.get_historial_contacto(nombre, telefono)
    return jsonify(historial)


@app.route('/api/pedidos/envios', methods=['GET'])
def listar_pedidos_envios():
    usuario = obtener_usuario_actual()
    if not usuario or usuario.rol != 'admin':
        return jsonify({'error': 'Acceso denegado.'}), 403

    fecha = request.args.get('fecha') or None
    # Now returns all envios including Entregado so they remain visible
    pedidos = models.get_all_pedidos(fecha=fecha, tipo_entrega='envio')
    
    # Sort: pending first, Entregado last
    pedidos_sorted = sorted(pedidos, key=lambda x: 1 if x['estado'] == 'Entregado' else 0)
    
    return jsonify(pedidos_sorted)


# ── API de estadísticas y configuración ─────────────────────────────────────

@app.route('/api/stats', methods=['GET'])
def estadisticas():
    usuario = obtener_usuario_actual()
    if not usuario or usuario.rol != 'admin':
        return jsonify({'error': 'Acceso denegado.'}), 403
    fecha = request.args.get('fecha') or None
    return jsonify(models.get_stats(fecha=fecha))


@app.route('/api/precios', methods=['GET'])
def precios():
    """Expone los precios al frontend para el cálculo en tiempo real."""
    return jsonify({
        'locro_unitario': PRECIO_LOCRO_UNITARIO,
        'pastelito_docena': PRECIO_PASTELITO_DOCENA,
        'pastelito_media_docena': PRECIO_PASTELITO_MEDIA_DOCENA,
        'pastelito_unidad': PRECIO_PASTELITO_UNIDAD,
    })


# ── Exportar a Excel ─────────────────────────────────────────────────────────

@app.route('/api/export')
def exportar_excel():
    usuario = obtener_usuario_actual()
    if not usuario or usuario.rol != 'admin':
        return jsonify({'error': 'Acceso denegado.'}), 403

    estado     = request.args.get('estado') or None
    medio_pago = request.args.get('medio_pago') or None
    fecha      = request.args.get('fecha') or None
    busqueda   = request.args.get('q') or None
    tipo_entrega = request.args.get('tipo_entrega') or None
    pedidos    = models.get_all_pedidos(estado=estado, medio_pago=medio_pago, fecha=fecha, busqueda=busqueda, tipo_entrega=tipo_entrega)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pedidos'

    # Habilitar líneas de cuadrícula visibles
    ws.views.sheetView[0].showGridLines = True

    encabezados = [
        'ID', 'Fecha', 'Cliente', 'Teléfono', 'Email', 'Dirección',
        'Locro (porciones)', 'Pastelitos Batata (unidades)', 'Pastelitos Membrillo (unidades)',
        'Medio de pago', 'Total ($)', 'Tipo entrega', 'Horario entrega', 'Notas', 'Estado', 'Pagado',
    ]

    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, titulo in enumerate(encabezados, start=1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Estilos de celdas
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    font_body = Font(name="Segoe UI", size=11)

    for row_idx, p in enumerate(pedidos, start=2):
        c_locro = sum(item['cantidad'] for item in p.get('items', []) if 'locro' in item.get('producto_nombre', '').lower())
        c_batata = sum(item['cantidad'] for item in p.get('items', []) if 'batata' in item.get('producto_nombre', '').lower())
        c_membrillo = sum(item['cantidad'] for item in p.get('items', []) if 'membrillo' in item.get('producto_nombre', '').lower())

        row_values = [
            p['id'],
            p['fecha_pedido'],
            p['nombre_cliente'],
            p['telefono'],
            p.get('email') or '',
            p['direccion'],
            c_locro,
            c_batata,
            c_membrillo,
            p['medio_pago'],
            p['monto_total'],
            'Retiro iglesia' if p.get('tipo_entrega') == 'retiro' else 'Envío domicilio',
            p.get('horario_entrega') or '',
            p.get('notas') or '',
            p['estado'],
            'Sí' if p.get('pagado') else 'No',
        ]
        
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            
            # Formatos de número y alineación
            if col_idx == 11:  # Total ($)
                cell.number_format = '$#,##0'
                cell.alignment = align_right
            elif col_idx in [1, 7, 8, 9]:  # ID y cantidades
                cell.alignment = align_right
            elif col_idx in [2, 4, 10, 12, 13, 15, 16]:  # Fecha, teléfono, pago, tipo entrega, horario, estado, pagado
                cell.alignment = align_center
            else:  # Cliente, dirección, email, notas
                cell.alignment = align_left

    # Auto-ajustar Anchos de Columnas Dinámicamente
    for col_idx, titulo in enumerate(encabezados, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(titulo))
        for row in range(2, len(pedidos) + 2):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                if col_idx == 11 and isinstance(val, (int, float)):
                    formatted_val = f"${val:,.0f}".replace(",", ".")
                    max_len = max(max_len, len(formatted_val))
                else:
                    max_len = max(max_len, len(str(val)))
        
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='pedidos_ventasRF.xlsx',
    )


# ── Rutas de Autenticación ───────────────────────────────────────────────────

@app.route('/api/auth/register', methods=['POST'])
def register_user():
    data = request.get_json() or {}
    nombre = data.get('nombre') or data.get('name')
    email = data.get('email')
    password = data.get('contrasenia') or data.get('password')
    codigo_staff = data.get('codigo_staff') or data.get('codigoStaff')
    
    if not nombre or not email or not password:
        return jsonify({'error': 'Nombre, email y contraseña son obligatorios.'}), 400
        
    nombre = str(nombre).strip()
    email = str(email).strip().lower()
    
    # Validar formato de email
    import re
    if not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
        return jsonify({'error': 'El formato del email es inválido.'}), 400
        
    if len(str(password)) < 6:
        return jsonify({'error': 'La contraseña debe tener al menos 6 caracteres.'}), 400
        
    # Determinar rol según el código de staff
    rol = 'customer'
    if codigo_staff:
        codigo_staff = str(codigo_staff).strip()
        if codigo_staff == CODIGO_ADMIN or codigo_staff.upper() == CODIGO_ADMIN:
            rol = 'admin'
        else:
            return jsonify({'error': 'Código de Staff inválido.'}), 400

    # Verificar si el usuario ya existe
    existing_user = models.get_usuario_by_email(email)
    if existing_user:
        return jsonify({'error': 'El correo electrónico ya está registrado.'}), 400
        
    # Crear hash de la contraseña
    password_hash = generate_password_hash(str(password))
    
    try:
        user = models.create_usuario(nombre, email, password_hash, rol)
        token = generar_token(user['id'])
        return jsonify({
            'token': token,
            'user': user
        }), 201
    except Exception as e:
        return jsonify({'error': f'Error al registrar el usuario: {str(e)}'}), 500


@app.route('/api/auth/login', methods=['POST'])
def login_user():
    data = request.get_json() or {}
    email = data.get('email')
    password = data.get('contrasenia') or data.get('password')
    
    if not email or not password:
        return jsonify({'error': 'Email y contraseña son obligatorios.'}), 400
        
    email = str(email).strip().lower()
    
    # Buscar usuario
    user = models.get_usuario_by_email(email)
    if not user:
        return jsonify({'error': 'Credenciales inválidas.'}), 401
        
    # Verificar contraseña
    if not check_password_hash(user['password_hash'], str(password)):
        return jsonify({'error': 'Credenciales inválidas.'}), 401
        
    token = generar_token(user['id'])
    return jsonify({
        'token': token,
        'user': user
    }), 200


ALLOWED_CLIENT_IDS = [
    '470092085691-g3qhlkdmgu2gkrj2qt6o428ja146e7t8.apps.googleusercontent.com', # Web
    '470092085691-hgphb8kuueta7bs70a284af80skvk1vq.apps.googleusercontent.com', # Android
    '470092085691-3s97ong1eao7ja6ae329h99muj9hh8ca.apps.googleusercontent.com'  # iOS
]

@app.route('/api/auth/google', methods=['POST'])
def google_login_user():
    data = request.get_json() or {}
    id_token_str = data.get('idToken')
    if not id_token_str:
        return jsonify({'error': 'Token de Google faltante.'}), 400

    try:
        request_transport = google_requests.Request()
        idinfo = id_token.verify_oauth2_token(id_token_str, request_transport)
        
        # Verificar que el cliente de origen esté en el listado de permitidos
        if idinfo.get('aud') not in ALLOWED_CLIENT_IDS:
            return jsonify({'error': 'Token de Google no está autorizado para esta aplicación.'}), 401
            
        email = idinfo.get('email')
        nombre = idinfo.get('name') or idinfo.get('given_name') or 'Usuario de Google'
        
        if not email:
            return jsonify({'error': 'No se pudo obtener el correo de Google.'}), 400
            
        email = email.strip().lower()
        
        # Buscar o registrar usuario
        user = models.get_usuario_by_email(email)
        if not user:
            user = models.create_usuario(nombre, email, 'google_oauth', 'user')
            
        token = generar_token(user['id'])
        return jsonify({
            'token': token,
            'user': user
        }), 200
        
    except ValueError as e:
        return jsonify({'error': f'Token de Google inválido: {str(e)}'}), 401
    except Exception as e:
        return jsonify({'error': f'Error al autenticar con Google: {str(e)}'}), 500


@app.route('/api/auth/me', methods=['GET'])
def get_current_user():
    token = None
    if 'Authorization' in request.headers:
        auth_header = request.headers['Authorization']
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
            
    if not token:
        return jsonify({'error': 'Token de autenticación faltante.'}), 401
        
    user_id = obtener_usuario_desde_token(token)
    if user_id in ('token_expirado', 'token_invalido'):
        return jsonify({'error': 'Token inválido o expirado.'}), 401
        
    current_user = models.get_usuario_by_id(user_id)
    if not current_user:
        return jsonify({'error': 'Usuario no encontrado.'}), 401
        
    return jsonify(current_user)


@app.route('/api/auth/upgrade-role', methods=['POST'])
def upgrade_role():
    """Permite a un usuario cambiar su rol ingresando un código de staff."""
    # Verificar autenticación
    token = None
    if 'Authorization' in request.headers:
        auth_header = request.headers['Authorization']
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if not token:
        return jsonify({'error': 'Token de autenticación faltante.'}), 401
    
    user_id = obtener_usuario_desde_token(token)
    if user_id in ('token_expirado', 'token_invalido'):
        return jsonify({'error': 'Token inválido o expirado.'}), 401

    data = request.get_json() or {}
    codigo = data.get('codigo', '').strip()

    if not codigo:
        return jsonify({'error': 'Debes ingresar un código de Staff.'}), 400

    # Determinar el nuevo rol
    nuevo_rol = None
    if codigo == CODIGO_ADMIN:
        nuevo_rol = 'admin'
    else:
        return jsonify({'error': 'Código de Staff inválido.'}), 400

    # Actualizar el rol en la base de datos
    ok = models.update_usuario_rol(user_id, nuevo_rol)
    if not ok:
        return jsonify({'error': 'No se pudo actualizar el rol.'}), 500

    # Devolver el usuario actualizado
    updated_user = models.get_usuario_by_id(user_id)
    return jsonify({
        'ok': True,
        'user': updated_user,
        'message': f'Tu rol ha sido actualizado a {nuevo_rol}.'
    })


@app.route('/api/auth/profile', methods=['PUT'])
def update_profile():
    # Verificar autenticación
    token = None
    if 'Authorization' in request.headers:
        auth_header = request.headers['Authorization']
        if auth_header.startswith('Bearer '):
            token = auth_header.split(' ')[1]
    
    if not token:
        return jsonify({'error': 'Token de autenticación faltante.'}), 401
    
    user_id = obtener_usuario_desde_token(token)
    if user_id in ('token_expirado', 'token_invalido'):
        return jsonify({'error': 'Token inválido o expirado.'}), 401

    data = request.get_json() or {}
    ok = models.update_usuario_profile(user_id, data)
    
    if not ok:
        return jsonify({'error': 'No se pudo actualizar el perfil.'}), 500

    updated_user = models.get_usuario_by_id(user_id)
    return jsonify({
        'ok': True,
        'user': updated_user
    })


# ── Arranque ─────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    socketio.run(app, debug=True, host='0.0.0.0', port=8080)
